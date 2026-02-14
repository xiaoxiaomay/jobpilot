#!/usr/bin/env python3
"""
Excel Job Tracker — Module A
==============================
Generates a professionally formatted .xlsx job tracking spreadsheet
from ranked job data with:
  - Dashboard summary sheet
  - Detailed job listings with conditional formatting
  - Immigration tracking columns
  - Application status tracking

Usage:
    python excel_tracker.py --input jobs_ranked.csv --output job_tracker.xlsx
    python excel_tracker.py --input jobs_ranked.csv  # defaults to job_tracker_YYYYMMDD.xlsx
"""

import os
import sys
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.chart import BarChart, PieChart, Reference

TOOLKIT_DIR = os.path.dirname(os.path.abspath(__file__))


# ─── Style Definitions ───────────────────────────────────────────────
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="2B579A")
SUBHEADER_FILL = PatternFill("solid", fgColor="4472C4")
DATA_FONT = Font(name="Arial", size=10)
LINK_FONT = Font(name="Arial", size=10, color="0563C1", underline="single")
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="2B579A")
SUBTITLE_FONT = Font(name="Arial", size=11, bold=True, color="333333")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# Priority colors
FILL_HIGH = PatternFill("solid", fgColor="FCE4EC")    # Light red
FILL_MEDIUM = PatternFill("solid", fgColor="FFF8E1")   # Light yellow
FILL_LOW = PatternFill("solid", fgColor="F5F5F5")      # Light gray

# Status colors
STATUS_FILLS = {
    "New": PatternFill("solid", fgColor="E3F2FD"),
    "Applied": PatternFill("solid", fgColor="FFF3E0"),
    "Interview": PatternFill("solid", fgColor="E8F5E9"),
    "Rejected": PatternFill("solid", fgColor="FFEBEE"),
    "Offer": PatternFill("solid", fgColor="C8E6C9"),
}

# Score color scale
SCORE_FILLS = {
    "high": PatternFill("solid", fgColor="C8E6C9"),    # Green
    "medium": PatternFill("solid", fgColor="FFF9C4"),   # Yellow
    "low": PatternFill("solid", fgColor="FFCDD2"),      # Red
}


def create_tracker(df, output_path):
    """Create the formatted Excel tracker."""
    wb = Workbook()
    
    # ─── Sheet 1: Dashboard ──────────────────────────────────────────
    create_dashboard(wb, df)
    
    # ─── Sheet 2: All Jobs ───────────────────────────────────────────
    create_jobs_sheet(wb, df)
    
    # ─── Sheet 3: Application Log ────────────────────────────────────
    create_application_log(wb, df)
    
    # ─── Sheet 4: Weekly Targets ─────────────────────────────────────
    create_weekly_targets(wb)
    
    wb.save(output_path)
    print(f"✅ Excel tracker saved: {output_path}")


def create_dashboard(wb, df):
    """Create the dashboard summary sheet."""
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_properties.tabColor = "2B579A"
    
    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "📊 JOB SEARCH TRACKER — DASHBOARD"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="2B579A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35
    
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Last updated: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Target: Data Scientist / AI Consultant / Product Analyst / Data Analyst"
    ws["A2"].font = Font(name="Arial", size=9, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")
    
    # ─── KPI Cards ────────────────────────────────────────────────
    row = 4
    ws[f"A{row}"] = "KEY METRICS"
    ws[f"A{row}"].font = SUBTITLE_FONT
    row += 1
    
    metrics = [
        ("Total Jobs Found", len(df)),
        ("🔴 High Priority", len(df[df["priority"].str.contains("HIGH", na=False)]) if "priority" in df.columns else 0),
        ("🟡 Medium Priority", len(df[df["priority"].str.contains("MEDIUM", na=False)]) if "priority" in df.columns else 0),
        ("BC PNP Tech Eligible", len(df[df["bcpnp_tech_eligible"].str.contains("Yes", na=False)]) if "bcpnp_tech_eligible" in df.columns else 0),
        ("Avg Skills Match", f"{df['score_skills'].mean():.0f}%" if "score_skills" in df.columns else "N/A"),
        ("Avg Total Score", f"{df['total_score'].mean():.0f}" if "total_score" in df.columns else "N/A"),
    ]
    
    for i, (label, value) in enumerate(metrics):
        col = (i % 3) * 3 + 1
        r = row + (i // 3) * 3
        
        cell_label = ws.cell(row=r, column=col, value=label)
        cell_label.font = Font(name="Arial", size=9, color="666666")
        
        cell_value = ws.cell(row=r + 1, column=col, value=value)
        cell_value.font = Font(name="Arial", size=20, bold=True, color="2B579A")
        cell_value.alignment = Alignment(horizontal="left")
    
    # ─── Score Distribution ───────────────────────────────────────
    row = 12
    ws[f"A{row}"] = "SCORE DISTRIBUTION"
    ws[f"A{row}"].font = SUBTITLE_FONT
    row += 1
    
    if "total_score" in df.columns:
        brackets = [
            ("75-100 (High)", len(df[df["total_score"] >= 75])),
            ("55-74 (Medium)", len(df[(df["total_score"] >= 55) & (df["total_score"] < 75)])),
            ("0-54 (Low)", len(df[df["total_score"] < 55])),
        ]
        
        ws.cell(row=row, column=1, value="Score Range").font = HEADER_FONT
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=2, value="Count").font = HEADER_FONT
        ws.cell(row=row, column=2).fill = HEADER_FILL
        
        for i, (label, count) in enumerate(brackets, 1):
            ws.cell(row=row + i, column=1, value=label).font = DATA_FONT
            ws.cell(row=row + i, column=2, value=count).font = DATA_FONT
    
    # ─── Top 10 Quick View ────────────────────────────────────────
    row = 18
    ws[f"A{row}"] = "TOP 10 JOBS — QUICK VIEW"
    ws[f"A{row}"].font = SUBTITLE_FONT
    row += 1
    
    quick_headers = ["#", "Score", "Title", "Company", "Skills", "Immigration", "BC PNP"]
    for i, h in enumerate(quick_headers):
        cell = ws.cell(row=row, column=i + 1, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
    
    for idx, (_, job) in enumerate(df.head(10).iterrows()):
        r = row + idx + 1
        ws.cell(row=r, column=1, value=idx + 1).font = DATA_FONT
        ws.cell(row=r, column=2, value=round(job.get("total_score", 0))).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=r, column=3, value=str(job.get("title", ""))[:50]).font = DATA_FONT
        ws.cell(row=r, column=4, value=str(job.get("company", ""))[:30]).font = DATA_FONT
        ws.cell(row=r, column=5, value=job.get("score_skills", "")).font = DATA_FONT
        ws.cell(row=r, column=6, value=job.get("score_immigration", "")).font = DATA_FONT
        ws.cell(row=r, column=7, value=str(job.get("bcpnp_tech_eligible", ""))).font = DATA_FONT
        
        # Color by priority
        score = job.get("total_score", 0)
        fill = FILL_HIGH if score >= 75 else (FILL_MEDIUM if score >= 55 else FILL_LOW)
        for c in range(1, 8):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).border = THIN_BORDER
    
    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14


def create_jobs_sheet(wb, df):
    """Create detailed job listings sheet."""
    ws = wb.create_sheet("All Jobs")
    ws.sheet_properties.tabColor = "4472C4"
    
    # Select and order columns for display
    display_cols = [
        "total_score", "priority", "title", "company", "location",
        "score_skills", "score_immigration", "score_salary", "score_company", "score_success",
        "bcpnp_tech_eligible", "noc_code_guess",
        "min_amount", "max_amount", "interval",
        "application_status", "notes",
        "job_url", "scraped_date", "search_query",
    ]
    
    available_cols = [c for c in display_cols if c in df.columns]
    df_display = df[available_cols].copy()
    
    # Headers
    col_labels = {
        "total_score": "Score",
        "priority": "Priority",
        "title": "Job Title",
        "company": "Company",
        "location": "Location",
        "score_skills": "Skills Match",
        "score_immigration": "Immigration Fit",
        "score_salary": "Salary Score",
        "score_company": "Company Score",
        "score_success": "Success Prob.",
        "bcpnp_tech_eligible": "BC PNP Tech",
        "noc_code_guess": "NOC Code",
        "min_amount": "Min Salary",
        "max_amount": "Max Salary",
        "interval": "Pay Period",
        "application_status": "Status",
        "notes": "Notes",
        "job_url": "URL",
        "scraped_date": "Found Date",
        "search_query": "Search Query",
    }
    
    # Write headers
    for col_idx, col_name in enumerate(available_cols, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_labels.get(col_name, col_name))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Write data
    for row_idx, (_, row) in enumerate(df_display.iterrows(), 2):
        for col_idx, col_name in enumerate(available_cols, 1):
            val = row.get(col_name)
            if pd.isna(val):
                val = ""
            
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(col_name in ["title", "notes"]))
            
            # URL as hyperlink
            if col_name == "job_url" and val and str(val).startswith("http"):
                cell.font = LINK_FONT
                try:
                    cell.hyperlink = str(val)
                except:
                    pass
            
            # Score color coding
            if col_name in ["score_skills", "score_immigration", "score_salary", "score_company", "score_success"]:
                try:
                    score_val = float(val) if val else 0
                    if score_val >= 70:
                        cell.fill = SCORE_FILLS["high"]
                    elif score_val >= 50:
                        cell.fill = SCORE_FILLS["medium"]
                    else:
                        cell.fill = SCORE_FILLS["low"]
                except (ValueError, TypeError):
                    pass
            
            # Priority row coloring
            if col_name == "priority":
                if "HIGH" in str(val):
                    cell.fill = FILL_HIGH
                elif "MEDIUM" in str(val):
                    cell.fill = FILL_MEDIUM
    
    # Auto-fit column widths
    col_widths = {
        "total_score": 8, "priority": 12, "title": 45, "company": 25,
        "location": 20, "score_skills": 12, "score_immigration": 14,
        "score_salary": 12, "score_company": 12, "score_success": 12,
        "bcpnp_tech_eligible": 12, "noc_code_guess": 22,
        "min_amount": 12, "max_amount": 12, "interval": 10,
        "application_status": 12, "notes": 30, "job_url": 50,
        "scraped_date": 12, "search_query": 20,
    }
    
    for col_idx, col_name in enumerate(available_cols, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(col_name, 15)
    
    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(available_cols))}{len(df_display) + 1}"


def create_application_log(wb, df):
    """Create application tracking log sheet."""
    ws = wb.create_sheet("Application Log")
    ws.sheet_properties.tabColor = "70AD47"
    
    headers = [
        "Date Applied", "Company", "Position", "Job URL",
        "Resume Version", "Cover Letter Version",
        "ATS Score", "Status", "Follow-up Date",
        "Response Date", "Interviewer", "Interview Notes",
        "Offer Details", "Decision"
    ]
    
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor="70AD47")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
    
    ws.freeze_panes = "A2"
    
    # Pre-populate with HIGH priority jobs
    high_priority = df[df.get("priority", pd.Series()).str.contains("HIGH", na=False)]
    for idx, (_, job) in enumerate(high_priority.head(20).iterrows(), 2):
        ws.cell(row=idx, column=2, value=str(job.get("company", ""))).font = DATA_FONT
        ws.cell(row=idx, column=3, value=str(job.get("title", ""))).font = DATA_FONT
        url = str(job.get("job_url", ""))
        if url.startswith("http"):
            ws.cell(row=idx, column=4, value=url).font = LINK_FONT
        ws.cell(row=idx, column=7, value=job.get("total_score", "")).font = DATA_FONT
        ws.cell(row=idx, column=8, value="To Apply").font = DATA_FONT
        
        for c in range(1, 15):
            ws.cell(row=idx, column=c).border = THIN_BORDER
    
    col_widths = [12, 25, 40, 50, 20, 20, 10, 12, 14, 14, 20, 40, 30, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def create_weekly_targets(wb):
    """Create weekly target tracking sheet."""
    ws = wb.create_sheet("Weekly Targets")
    ws.sheet_properties.tabColor = "ED7D31"
    
    ws.merge_cells("A1:F1")
    ws["A1"] = "📅 WEEKLY JOB SEARCH TARGETS"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    
    # Target metrics
    row = 3
    ws[f"A{row}"] = "Weekly Targets"
    ws[f"A{row}"].font = SUBTITLE_FONT
    row += 1
    
    targets = [
        ("Jobs reviewed", "50+", "Review scraped jobs and update rankings"),
        ("Applications submitted", "10-15", "Customized resume + cover letter per application"),
        ("Networking events/messages", "5+", "LinkedIn messages, meetups, coffee chats"),
        ("ATS score target", "≥70%", "Don't submit below 65%"),
        ("GitHub commits", "5+", "Portfolio projects, contributions"),
        ("LinkedIn activity", "3+", "Posts, comments, shares"),
    ]
    
    headers = ["Metric", "Target", "Notes", "Mon", "Tue", "Wed", "Thu", "Fri", "Total"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor="ED7D31")
        cell.border = THIN_BORDER
    
    for idx, (metric, target, notes) in enumerate(targets, 1):
        r = row + idx
        ws.cell(row=r, column=1, value=metric).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row=r, column=2, value=target).font = DATA_FONT
        ws.cell(row=r, column=3, value=notes).font = Font(name="Arial", size=9, color="666666")
        # Total formula
        ws.cell(row=r, column=9, value=f"=SUM(D{r}:H{r})").font = Font(name="Arial", size=10, bold=True)
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = THIN_BORDER
    
    col_widths = [25, 10, 45, 8, 8, 8, 8, 8, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Excel Job Tracker Generator")
    parser.add_argument("--input", required=True, help="Ranked CSV from job_ranker.py")
    parser.add_argument("--output", default=None, help="Output .xlsx path")
    args = parser.parse_args()
    
    if args.output is None:
        date_str = datetime.now().strftime("%Y%m%d")
        args.output = f"job_tracker_{date_str}.xlsx"
    
    print("=" * 60)
    print("EXCEL JOB TRACKER GENERATOR")
    print("=" * 60)
    
    df = pd.read_csv(args.input)
    print(f"Loaded {len(df)} ranked jobs")
    
    create_tracker(df, args.output)
    
    print(f"\n📊 Sheets created:")
    print(f"  1. Dashboard — KPIs and top 10 quick view")
    print(f"  2. All Jobs — Full listing with scores and filters")
    print(f"  3. Application Log — Track your applications")
    print(f"  4. Weekly Targets — Job search activity tracking")


if __name__ == "__main__":
    main()
